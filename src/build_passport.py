import json
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))
from boq_items import ITEMS
from carbon_data import lookup_carbon

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE = os.path.join(ROOT, "AMP_Passport_Template.xlsx")
OUT_DIR = os.path.join(ROOT, "output")

UNIT_MAP = {
    "cu.m": "cum", "cum": "cum", "m³": "cum",
    "sq.m": "sqm", "sqm": "sqm", "m²": "sqm",
    "mtr.": "m", "mtr": "m", "m": "m",
    "kg.": "kg", "kg": "kg",
    "each": "nos", "nos": "nos",
}


def normalize_unit(raw_unit: str):
    u = (raw_unit or "").strip().lower()
    if u == "100 sq.m":
        return "sqm", 100.0, "Unit printed as '100 Sq.m' (DSR per-100-sqm convention); quantity multiplied by 100"
    if u == "10 cubic decimetre":
        return "cum", 0.01, "Unit printed as '10 Cubic decimetre' = 0.01 cum per Instructions sheet"
    return UNIT_MAP.get(u, u), 1.0, ""


def quantity_column(unit: str):
    return {"cum": "volume", "sqm": "area", "m": "length", "kg": "weight", "nos": "count"}.get(unit)


AREA_TO_MASS = {
    "9": ("thickness_m", 0.040),
    "10": ("kg_per_m2", 1.7),
    "21": ("thickness_m", 0.070),
    "22": ("thickness_m", 0.115),
    "23": ("thickness_m", 0.115),
    "25": ("thickness_m", 0.035),
    "26": ("thickness_m", 0.025),
    "40": ("thickness_m", 0.040),
    "41": ("thickness_m", 0.018),
    "42": ("thickness_m", 0.040),
    "44": ("kg_per_m2", 1.7),
    "52": ("thickness_m", 0.012),
    "53": ("thickness_m", 0.015),
    "54": ("thickness_m", 0.006),
    "55": ("thickness_m", 0.018),
    "64": ("thickness_m", 0.050),
}
LENGTH_TO_MASS = {
    "43": ("cross_section_m2", 0.040 * 0.006),
    "47": ("cross_section_m2", 0.15 * 0.15),
}
COUNT_TO_MASS = {
    "48": ("volume_per_unit_m3", 0.45 * 0.45 * 0.05),
}


def derive_mass_kg(item_no, qcol, qty_norm, carbon):
    if qty_norm is None or carbon is None:
        return None, None
    if qcol == "area" and item_no in AREA_TO_MASS:
        kind, value = AREA_TO_MASS[item_no]
        if kind == "thickness_m":
            return (qty_norm * value * carbon["density"],
                    f"Mass derived from area x {value * 1000:.0f}mm thickness (from BoQ description) x density")
        if kind == "kg_per_m2":
            return qty_norm * value, f"Mass derived from DSR-stated coating rate {value} kg/sq.m"
    if qcol == "length" and item_no in LENGTH_TO_MASS:
        _, value = LENGTH_TO_MASS[item_no]
        return (qty_norm * value * carbon["density"],
                f"Mass derived from length x {value * 1e6:.1f} sq.mm cross-section (from BoQ description) x density")
    if qcol == "count" and item_no in COUNT_TO_MASS:
        _, value = COUNT_TO_MASS[item_no]
        return (qty_norm * value * carbon["density"],
                f"Mass derived from count x {value * 1e6:.1f} cu.cm volume per unit (from BoQ description) x density")
    return None, None


def _row_comment(item, e, unit_note, carbon, derived_mass_note):
    parts = []
    if item["excluded"]:
        parts.append(f"[EXCLUDED] {item['excluded']}")
    if unit_note:
        parts.append(unit_note)
    if item["no"] == "17" and e.get("suffix") == "ii":
        parts.append("Scan shows alternate qty 1500.0 Kg marked * for Seismic Zone V; "
                      "1375.0 Kg used as base figure (ambiguous which value the * applies to)")
    if item["no"] in ("19", "20", "21", "22", "23"):
        parts.append("Brick class designation left blank in source BoQ per footnote "
                      "('appropriate class designation of bricks may be incorporated "
                      "before calling tenders') - not a scan defect")
    if item["no"] in ("24", "25"):
        parts.append("Wood species left blank in source BoQ per the same footnote")
    if e.get("code") in ("N.S.I.",):
        parts.append("N.S.I. = Non-Schedule Item (rate outside DSR 1989)")
    if carbon:
        parts.append(f"Embodied carbon source: {carbon['source']}")
    if derived_mass_note:
        parts.append(derived_mass_note)
    if item["no"] in ("22", "23") and derived_mass_note:
        parts.append("Half-brick thickness assumed at standard 115mm nominal (not stated on scan)")
    if item["no"] == "47" and derived_mass_note:
        parts.append("Gola cross-section treated as full 15x15cm square; actual filleted "
                      "profile is smaller, so mass is an upper-bound estimate")
    if item["no"] in ("42", "55") and derived_mass_note:
        parts.append("Composite assembly: mass uses only the top finish layer's stated "
                      "thickness/density; underlayer(s) of a different material are excluded")
    return "; ".join(parts)


def _embodied_carbon_kg(qcol, carbon, qty_norm, derived_mass_kg, weight_kg):
    if not carbon:
        return None
    if weight_kg is not None:
        return round(weight_kg * carbon["gwp_per_kg"], 1)
    if qty_norm is None:
        return None
    mass_kg = qty_norm * carbon["density"] if qcol == "volume" else derived_mass_kg
    return round(mass_kg * carbon["gwp_per_kg"], 1) if mass_kg else None


def build_rows():
    rows = []
    gmap_seq = 1
    for item in ITEMS:
        entries = item["sub"] if item["sub"] else [dict(suffix="", label="", qty=item["qty"],
                                                          unit=item["unit"], code=item["code"])]
        for e in entries:
            boq_no = f"{item['no']}.{e['suffix']}" if e["suffix"] else item["no"]
            desc = item["desc"] if not e["label"] else f"{item['desc']} :: {e['suffix']}) {e['label']}"
            unit_norm, mult, unit_note = normalize_unit(e["unit"])
            qty_raw = e["qty"]
            qty_norm = round(qty_raw * mult, 4) if qty_raw is not None else None
            qcol = quantity_column(unit_norm) if qty_norm is not None else None

            carbon = None if item["excluded"] else lookup_carbon(item["material"])
            derived_mass_kg, derived_mass_note = derive_mass_kg(item["no"], qcol, qty_norm, carbon)
            comment = _row_comment(item, e, unit_note, carbon, derived_mass_note)

            row = {
                "GMAP Id": f"GMAP-{gmap_seq:03d}",
                "BOQ Item No.": boq_no,
                "Article Number": "",
                "External DB Id": "",
                "Description": desc,
                "Floor / Section": item["section"],
                "Discipline": item["discipline"],
                "Material / Product": item["material"],
                "All Materials Detected": item["material"],
                "Material Category": item["category"],
                "Material Confidence": "Medium" if item["excluded"] else "High",
                "Grade": item["grade"],
                "Mix Ratio": item["grade"],
                "Original Quantity": qty_raw,
                "Original Unit": e["unit"],
                "Volume (m³)": qty_norm if qcol == "volume" else None,
                "Area (m²)": qty_norm if qcol == "area" else None,
                "Length (m)": qty_norm if qcol == "length" else None,
                "Weight (kg)": qty_norm if qcol == "weight" else None,
                "Count (Nos)": qty_norm if qcol == "count" else None,
                "Derived Quantity": qty_norm if unit_note else None,
                "Derived Quantity Unit": unit_norm if unit_note else "",
                "Derived Quantity Basis": unit_note,
                "Density (kg/m³)": carbon["density"] if carbon else None,
                "Embodied Carbon A1-A3 (kg CO₂e)": None,
                "GWP / kg (kg CO₂e/kg)": carbon["gwp_per_kg"] if carbon else None,
                "Schedule (DSR/SOR)": "DSR 1989",
                "Schedule Item Code": e.get("code", ""),
                "Standard / Code Reference": "",
                "Classification (Matched)": "",
                "% Reused": None, "% Available for Reuse": None, "Assumed Construction Waste": None,
                "Waste Codes": "", "Detachability – Connection": "", "Detachability – Connection Detail": "",
                "Detachability – Accessibility": "", "Detachability – Intersection": "",
                "Detachability – Product Edge": "", "Lifespan (Years)": None,
                "Length (mm)": None, "Width (mm)": None, "Height (mm)": None, "Thickness (mm)": None,
                "Depth (mm)": None, "Diameter (mm)": None,
                "Unit Rate": None, "Total Cost": None, "Currency": "INR",
                "Comment": comment,
            }
            row["Embodied Carbon A1-A3 (kg CO₂e)"] = _embodied_carbon_kg(
                qcol, carbon, qty_norm, derived_mass_kg, row["Weight (kg)"])

            rows.append(row)
            gmap_seq += 1
    return rows


def write_xlsx(rows):
    wb = load_workbook(TEMPLATE)
    ws = wb["Material Passport"]
    headers = [ws.cell(3, c).value for c in range(1, 51)]
    for i, row in enumerate(rows):
        r = 7 + i
        for c, h in enumerate(headers, start=1):
            ws.cell(r, c, row.get(h))
    out_path = os.path.join(OUT_DIR, "passport_filled.xlsx")
    wb.save(out_path)
    return out_path


def write_json(rows):
    out_path = os.path.join(OUT_DIR, "passport.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    return out_path


def write_building_meta():
    meta = {
        "name_of_work": "Construction of Principal's Residence (General-Modified), CBRI Roorkee",
        "location": "Central Building Research Institute (CBRI), Roorkee, Uttarakhand, India",
        "schedule": "Schedule 'A'",
        "depth_of_foundation_m": 0.60,
        "plinth_height_m": 0.45,
        "plinth_area_sqm": 90.6,
        "no_of_boq_items": 64,
        "seismic_zone": "Zone I-IV (base design); Zone V alternate reinforcement noted for item 17.ii",
        "safe_bearing_capacity": "10 T/Sq.m and above",
        "class_designation_of_soil": None,
        "class_designation_of_soil_note": "Illegible - physically cut off at the bottom edge of the scanned "
                                           "page 1 (confirmed via 400 DPI crop render, not a rendering artifact)",
        "rate_schedule": "DSR 1989 (Delhi Schedule of Rates)",
        "source_document": "BoQ_CBRI_Principals_Residence.pdf",
    }
    out_path = os.path.join(OUT_DIR, "building_meta.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)
    return out_path


def write_visualization(rows):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from collections import Counter

    counts = Counter(r["Material Category"] for r in rows if not r["Comment"].startswith("[EXCLUDED]"))
    categories = sorted(counts, key=lambda k: -counts[k])
    values = [counts[k] for k in categories]

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(categories, values, color="#2b7a78")
    ax.set_xlabel("Number of BoQ line items")
    ax.set_title("Material Category Distribution — CBRI Principal's Residence BoQ")
    ax.invert_yaxis()
    for i, v in enumerate(values):
        ax.text(v + 0.1, i, str(v), va="center")
    fig.tight_layout()
    out_path = os.path.join(OUT_DIR, "visualization.png")
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    rows = build_rows()
    xlsx_path = write_xlsx(rows)
    json_path = write_json(rows)
    meta_path = write_building_meta()
    viz_path = write_visualization(rows)
    print(f"Rows built: {len(rows)}")
    print(f"Wrote: {xlsx_path}")
    print(f"Wrote: {json_path}")
    print(f"Wrote: {meta_path}")
    print(f"Wrote: {viz_path}")


if __name__ == "__main__":
    main()
